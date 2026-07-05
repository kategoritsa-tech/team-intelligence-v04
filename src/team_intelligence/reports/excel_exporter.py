from datetime import date, datetime, time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from team_intelligence.analytics.activity_analyzer import ActivityAnalyzer
from team_intelligence.daily.models import AuthorDailyMetrics, DailyReportMetrics, StructuredDailyReport


class ExcelExporter:
    """Экспорт аналитики в Excel."""

    HEADER_FILL = PatternFill(fill_type="solid", fgColor="4472C4")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")

    def __init__(self, output_dir: Path):
        self.output_dir = output_dir
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def export_summary(
        self,
        analyzer,
        structured_reports: list[StructuredDailyReport] | None = None,
        report_metrics: list[DailyReportMetrics] | None = None,
        author_daily_metrics: list[AuthorDailyMetrics] | None = None,
    ) -> Path:
        """Создает Excel-отчет."""

        workbook = Workbook()

        self.create_summary_sheet(workbook, analyzer, structured_reports, report_metrics)
        self.create_authors_sheet(workbook, analyzer)
        self.create_activity_sheet(workbook, analyzer)
        self.create_messages_sheet(workbook, analyzer)

        if structured_reports is not None:
            self.create_daily_reports_sheet(workbook, structured_reports, report_metrics or [])

        if structured_reports is not None:
            self.create_executive_summary_sheet(workbook, structured_reports, report_metrics or [])

        if structured_reports is not None:
            self.create_daily_review_sheet(workbook, structured_reports, report_metrics or [])

        if structured_reports is not None:
            self.create_author_review_sheet(workbook, structured_reports, report_metrics or [])

        if structured_reports is not None:
            self.create_report_antipatterns_sheet(workbook, structured_reports, report_metrics or [])

        if author_daily_metrics is not None:
            self.create_daily_authors_sheet(workbook, author_daily_metrics)

        file_path = self.output_dir / "team_intelligence.xlsx"
        workbook.save(file_path)

        return file_path

    def create_summary_sheet(
        self,
        workbook,
        analyzer,
        structured_reports: list[StructuredDailyReport] | None = None,
        report_metrics: list[DailyReportMetrics] | None = None,
    ):
        """Создает лист Summary."""

        sheet = workbook.active
        sheet.title = "Summary"

        sheet.append(["Показатель", "Значение"])
        sheet.append(["💬 Количество сообщений", analyzer.total_messages()])
        sheet.append(["👥 Количество участников", analyzer.unique_authors()])
        sheet.append(["📅 Первый день", self.excel_safe_value(analyzer.first_message_date())])
        sheet.append(["📅 Последний день", self.excel_safe_value(analyzer.last_message_date())])
        sheet.append(["📆 Длительность переписки", f"{analyzer.days_count()} дней"])
        sheet.append(["📈 Среднее сообщений в день", analyzer.average_messages_per_day()])
        sheet.append(["🏆 Самый активный автор", analyzer.top_author()])

        if structured_reports is not None:
            sheet.append(["🧾 Найдено daily-отчетов", len(structured_reports)])

        if report_metrics:
            average_quality = round(
                sum(metric.quality_score for metric in report_metrics) / len(report_metrics),
                1,
            )
            average_completeness = round(
                sum(metric.completeness_score for metric in report_metrics) / len(report_metrics),
                1,
            )
            late_reports = sum(1 for metric in report_metrics if metric.is_late)
            sheet.append(["⭐ Среднее качество daily", average_quality])
            sheet.append(["✅ Средняя полнота daily", average_completeness])
            sheet.append(["⏰ Опоздавших отчетов", late_reports])

        self.format_sheet(sheet)

    def create_authors_sheet(self, workbook, analyzer):
        """Создает лист Authors."""

        sheet = workbook.create_sheet("Authors")
        sheet.append(["Автор", "Количество сообщений"])

        for author, count in analyzer.messages_by_author().most_common():
            sheet.append([author, count])

        self.format_sheet(sheet)

    def create_activity_sheet(self, workbook, analyzer):
        """Создает лист Activity."""

        sheet = workbook.create_sheet("Activity")
        sheet.append(["Дата", "Количество сообщений"])

        activity_analyzer = ActivityAnalyzer(analyzer.messages)

        for day, count in sorted(activity_analyzer.messages_by_day().items()):
            sheet.append([self.excel_safe_value(day), count])

        self.format_sheet(sheet)
        self.add_activity_chart(sheet)

    def add_activity_chart(self, sheet):
        """Добавляет график активности."""

        if sheet.max_row < 2:
            return

        chart = LineChart()
        chart.title = "Активность по дням"
        chart.style = 2
        chart.y_axis.title = "Сообщений"
        chart.x_axis.title = "Дата"

        values = Reference(sheet, min_col=2, min_row=1, max_row=sheet.max_row)
        dates = Reference(sheet, min_col=1, min_row=2, max_row=sheet.max_row)

        chart.add_data(values, titles_from_data=True, from_rows=False)
        chart.set_categories(dates)
        chart.width = 18
        chart.height = 8

        sheet.add_chart(chart, "D2")

    def create_messages_sheet(self, workbook, analyzer):
        """Создает лист со всеми сообщениями."""

        sheet = workbook.create_sheet("Messages")
        sheet.append(["Дата", "Автор", "Сообщение", "Вложение"])

        for message in analyzer.messages:
            sheet.append(
                [
                    self.excel_safe_value(message.date),
                    message.author,
                    message.text,
                    "Да" if message.has_media else "Нет",
                ]
            )

        self.format_sheet(sheet)

    def create_executive_summary_sheet(
        self,
        workbook,
        structured_reports: list[StructuredDailyReport],
        report_metrics: list[DailyReportMetrics],
    ):
        """Создает управленческую сводку по daily-отчетам."""

        sheet = workbook.create_sheet(
            "Executive Summary",
            0,
        )

        sheet.append(
            [
                "Блок",
                "Метрика",
                "Значение",
            ]
        )

        total_reports = len(structured_reports)
        total_authors = len(
            {
                report.author
                for report in structured_reports
            }
        )

        total_metrics = len(report_metrics)

        if total_metrics > 0:
            avg_quality = round(
                sum(metric.quality_score for metric in report_metrics) / total_metrics,
                1,
            )

            avg_completeness = round(
                sum(metric.completeness_score for metric in report_metrics) / total_metrics,
                1,
            )
        else:
            avg_quality = 0
            avg_completeness = 0

        problematic_reports = 0
        late_reports = 0
        reports_without_completed = 0
        reports_without_plan = 0
        reports_with_problems = 0

        antipatterns = {}

        def add_antipattern(name: str):
            antipatterns[name] = antipatterns.get(
                name,
                0,
            ) + 1

        for metric in report_metrics:
            is_problematic = False

            if metric.quality_score < 70:
                is_problematic = True
                add_antipattern("Низкое качество отчета")

            if metric.completeness_score < 80:
                is_problematic = True
                add_antipattern("Неполный отчет")

            if metric.is_late:
                is_problematic = True
                late_reports += 1
                add_antipattern("Отчет после дедлайна")

            if metric.completed_tasks_count == 0:
                is_problematic = True
                reports_without_completed += 1
                add_antipattern("Нет выполненных задач")

            if metric.planned_tasks_count == 0:
                is_problematic = True
                reports_without_plan += 1
                add_antipattern("Нет плана на сегодня")

            if metric.problems_count > 0:
                is_problematic = True
                reports_with_problems += 1
                add_antipattern("Есть проблемы или блокеры")

            if metric.notes:
                is_problematic = True

            for note in metric.notes:
                note_lower = note.lower()

                if "мало конкретики" in note_lower:
                    add_antipattern("Мало конкретики")

                if "jira" in note_lower or "джира" in note_lower:
                    add_antipattern("Нет привязки к задаче")

            if is_problematic:
                problematic_reports += 1

        metrics_by_author = {}

        for metric in report_metrics:
            metrics_by_author.setdefault(
                metric.author,
                [],
            ).append(metric)

        high_risk_authors = 0
        medium_risk_authors = 0

        for metrics in metrics_by_author.values():
            reports_count = len(metrics)

            if reports_count == 0:
                continue

            avg_author_quality = (
                sum(metric.quality_score for metric in metrics) / reports_count
            )

            avg_author_completeness = (
                sum(metric.completeness_score for metric in metrics) / reports_count
            )

            late_count = sum(
                1 for metric in metrics
                if metric.is_late
            )

            no_completed_count = sum(
                1 for metric in metrics
                if metric.completed_tasks_count == 0
            )

            no_plan_count = sum(
                1 for metric in metrics
                if metric.planned_tasks_count == 0
            )

            problems_count = sum(
                1 for metric in metrics
                if metric.problems_count > 0
            )

            notes_count = sum(
                1 for metric in metrics
                if metric.notes
            )

            risk_score = 0

            if avg_author_quality < 60:
                risk_score += 3
            elif avg_author_quality < 70:
                risk_score += 2

            if avg_author_completeness < 70:
                risk_score += 3
            elif avg_author_completeness < 80:
                risk_score += 2

            if no_completed_count >= max(2, reports_count // 2):
                risk_score += 3
            elif no_completed_count > 0:
                risk_score += 1

            if no_plan_count >= max(2, reports_count // 2):
                risk_score += 3
            elif no_plan_count > 0:
                risk_score += 1

            if problems_count >= max(2, reports_count // 2):
                risk_score += 2
            elif problems_count > 0:
                risk_score += 1

            if late_count >= max(2, reports_count // 2):
                risk_score += 2
            elif late_count > 0:
                risk_score += 1

            if notes_count >= max(2, reports_count // 2):
                risk_score += 2

            if risk_score >= 7:
                high_risk_authors += 1
            elif risk_score >= 3:
                medium_risk_authors += 1

        if antipatterns:
            top_antipattern = max(
                antipatterns.items(),
                key=lambda item: item[1],
            )
            top_antipattern_text = f"{top_antipattern[0]} ({top_antipattern[1]})"
        else:
            top_antipattern_text = "Не выявлено"

        problematic_percent = 0

        if total_metrics > 0:
            problematic_percent = round(
                problematic_reports / total_metrics * 100,
                1,
            )

        rows = [
            ["Общая статистика", "Всего отчетов", total_reports],
            ["Общая статистика", "Всего авторов", total_authors],
            ["Общая статистика", "Проблемных отчетов", problematic_reports],
            ["Общая статистика", "Доля проблемных отчетов", f"{problematic_percent}%"],

            ["Качество", "Среднее качество", avg_quality],
            ["Качество", "Средняя полнота", avg_completeness],

            ["Дисциплина", "Отчетов после дедлайна", late_reports],
            ["Дисциплина", "Отчетов без выполненных задач", reports_without_completed],
            ["Дисциплина", "Отчетов без плана", reports_without_plan],
            ["Дисциплина", "Отчетов с проблемами/блокерами", reports_with_problems],

            ["Риски", "Авторов высокого риска", high_risk_authors],
            ["Риски", "Авторов среднего риска", medium_risk_authors],

            ["Антипаттерны", "Самая частая проблема", top_antipattern_text],
        ]

        for row in rows:
            sheet.append(row)

        self.format_sheet(sheet)

    def create_daily_reports_sheet(
        self,
        workbook,
        structured_reports: list[StructuredDailyReport],
        report_metrics: list[DailyReportMetrics],
    ):
        """Создает лист с разбором daily-отчетов."""

        metrics_by_key = {
            (metric.author, metric.report_date): metric
            for metric in report_metrics
        }

        sheet = workbook.create_sheet("Daily Reports")
        sheet.append(
            [
                "Дата",
                "Автор",
                "Роль",
                "Время",
                "Вчера задач",
                "Сегодня задач",
                "Проблемы",
                "Jira",
                "Качество",
                "Полнота",
                "Опоздал",
                "Текст отчета",
                "Замечания",
            ]
        )

        for report in structured_reports:
            metric = metrics_by_key.get((report.author, report.report_date))
            sheet.append(
                [
                    self.excel_safe_value(report.report_date),
                    report.author,
                    report.role,
                    self.excel_safe_value(metric.post_time if metric else None),
                    report.completed_tasks_count,
                    report.planned_tasks_count,
                    len(report.problems) if report.has_problems else 0,
                    len(report.jira_keys),
                    metric.quality_score if metric else None,
                    metric.completeness_score if metric else None,
                    "Да" if metric and metric.is_late else "Нет",
                    report.raw_text,
                    "; ".join(metric.notes) if metric else "",
                ]
            )

        self.format_sheet(sheet)

    def create_daily_authors_sheet(
        self,
        workbook,
        author_metrics: list[AuthorDailyMetrics],
    ):
        """Создает лист со сводкой по сотрудникам."""

        sheet = workbook.create_sheet("Daily Authors")
        sheet.append(
            [
                "Автор",
                "Отчетов",
                "Среднее качество",
                "Средняя полнота",
                "Среднее сделано",
                "Среднее план",
                "Опозданий",
                "Отчетов с проблемами",
                "Уклончивых пунктов",
            ]
        )

        for metric in author_metrics:
            sheet.append(
                [
                    metric.author,
                    metric.reports_count,
                    metric.average_quality,
                    metric.average_completeness,
                    metric.average_completed_tasks,
                    metric.average_planned_tasks,
                    metric.late_reports,
                    metric.reports_with_problems,
                    metric.vague_items_count,
                ]
            )

        self.format_sheet(sheet)

    def create_report_antipatterns_sheet(
        self,
        workbook,
        structured_reports: list[StructuredDailyReport],
        report_metrics: list[DailyReportMetrics],
    ):
        """Создает лист с типовыми антипаттернами daily-отчетов."""

        reports_by_key = {
            (
                report.author,
                report.report_date,
            ): report
            for report in structured_reports
        }

        antipatterns = {}

        def add_antipattern(
            name: str,
            metric: DailyReportMetrics,
            detail: str,
        ):
            report = reports_by_key.get(
                (
                    metric.author,
                    metric.report_date,
                )
            )

            if name not in antipatterns:
                antipatterns[name] = {
                    "count": 0,
                    "authors": set(),
                    "details": [],
                    "examples": [],
                }

            item = antipatterns[name]
            item["count"] += 1
            item["authors"].add(metric.author)

            if detail and detail not in item["details"]:
                item["details"].append(detail)

            if report is not None and len(item["examples"]) < 5:
                text = report.raw_text.strip().replace("\n", " / ")

                if len(text) > 500:
                    text = text[:500] + "..."

                item["examples"].append(
                    f"{metric.author}, {self.excel_safe_value(metric.report_date)}: {text}"
                )

        for metric in report_metrics:
            if metric.completed_tasks_count == 0:
                add_antipattern(
                    "Нет выполненных задач",
                    metric,
                    "В отчете не указано, что было сделано за вчера.",
                )

            if metric.planned_tasks_count == 0:
                add_antipattern(
                    "Нет плана на сегодня",
                    metric,
                    "В отчете не указан план работ на сегодня.",
                )

            if metric.problems_count > 0:
                add_antipattern(
                    "Есть проблемы или блокеры",
                    metric,
                    "В отчете указаны проблемы, блокеры или ограничения.",
                )

            if metric.is_late:
                add_antipattern(
                    "Отчет после дедлайна",
                    metric,
                    "Отчет опубликован позже установленного времени.",
                )

            if metric.quality_score < 70:
                add_antipattern(
                    "Низкое качество отчета",
                    metric,
                    "Оценка качества отчета ниже 70.",
                )

            if metric.completeness_score < 80:
                add_antipattern(
                    "Неполный отчет",
                    metric,
                    "Оценка полноты отчета ниже 80.",
                )

            for note in metric.notes:
                note_lower = note.lower()

                if "мало конкретики" in note_lower:
                    add_antipattern(
                        "Мало конкретики",
                        metric,
                        "Формулировки слишком общие, сложно понять конкретный результат.",
                    )

                if "jira" in note_lower or "джира" in note_lower:
                    add_antipattern(
                        "Нет привязки к задаче",
                        metric,
                        "В отчете нет понятной ссылки на задачу, тикет или рабочий объект.",
                    )

        sheet = workbook.create_sheet("Report Antipatterns")

        sheet.append(
            [
                "Антипаттерн",
                "Кол-во отчетов",
                "Кол-во авторов",
                "Авторы",
                "Описание",
                "Примеры",
            ]
        )

        rows = []

        for name, data in antipatterns.items():
            rows.append(
                [
                    name,
                    data["count"],
                    len(data["authors"]),
                    ", ".join(sorted(data["authors"])),
                    "; ".join(data["details"]),
                    "\n\n".join(data["examples"]),
                ]
            )

        rows.sort(
            key=lambda row: row[1],
            reverse=True,
        )

        for row in rows:
            sheet.append(row)

        self.format_sheet(sheet)

    def create_author_review_sheet(
        self,
        workbook,
        structured_reports: list[StructuredDailyReport],
        report_metrics: list[DailyReportMetrics],
    ):
        """Создает лист с агрегированной оценкой авторов daily-отчетов."""

        reports_by_author = {}

        for report in structured_reports:
            reports_by_author.setdefault(
                report.author,
                [],
            ).append(report)

        metrics_by_author = {}

        for metric in report_metrics:
            metrics_by_author.setdefault(
                metric.author,
                [],
            ).append(metric)

        rows = []

        for author, metrics in metrics_by_author.items():
            reports = reports_by_author.get(
                author,
                [],
            )

            role = ""

            for report in reports:
                if report.role:
                    role = report.role
                    break

            reports_count = len(metrics)

            if reports_count == 0:
                continue

            avg_quality = round(
                sum(metric.quality_score for metric in metrics) / reports_count,
                1,
            )

            avg_completeness = round(
                sum(metric.completeness_score for metric in metrics) / reports_count,
                1,
            )

            late_count = sum(
                1 for metric in metrics
                if metric.is_late
            )

            no_completed_count = sum(
                1 for metric in metrics
                if metric.completed_tasks_count == 0
            )

            no_plan_count = sum(
                1 for metric in metrics
                if metric.planned_tasks_count == 0
            )

            problems_count = sum(
                1 for metric in metrics
                if metric.problems_count > 0
            )

            notes_count = sum(
                1 for metric in metrics
                if metric.notes
            )

            reasons = []
            risk_score = 0

            if avg_quality < 60:
                reasons.append("Среднее качество ниже 60")
                risk_score += 3
            elif avg_quality < 70:
                reasons.append("Среднее качество ниже 70")
                risk_score += 2

            if avg_completeness < 70:
                reasons.append("Средняя полнота ниже 70")
                risk_score += 3
            elif avg_completeness < 80:
                reasons.append("Средняя полнота ниже 80")
                risk_score += 2

            if no_completed_count >= max(2, reports_count // 2):
                reasons.append("Часто нет выполненных задач")
                risk_score += 3
            elif no_completed_count > 0:
                reasons.append("Есть отчеты без выполненных задач")
                risk_score += 1

            if no_plan_count >= max(2, reports_count // 2):
                reasons.append("Часто нет плана")
                risk_score += 3
            elif no_plan_count > 0:
                reasons.append("Есть отчеты без плана")
                risk_score += 1

            if problems_count >= max(2, reports_count // 2):
                reasons.append("Часто есть проблемы или блокеры")
                risk_score += 2
            elif problems_count > 0:
                reasons.append("Есть проблемы или блокеры")
                risk_score += 1

            if late_count >= max(2, reports_count // 2):
                reasons.append("Часто опаздывает с отчетами")
                risk_score += 2
            elif late_count > 0:
                reasons.append("Есть опоздания")
                risk_score += 1

            if notes_count >= max(2, reports_count // 2):
                reasons.append("Много отчетов с замечаниями")
                risk_score += 2

            if risk_score >= 7:
                risk = "Высокий"
                priority = 1
            elif risk_score >= 3:
                risk = "Средний"
                priority = 2
            else:
                risk = "Низкий"
                priority = 3

            rows.append(
                [
                    priority,
                    author,
                    role,
                    reports_count,
                    avg_quality,
                    avg_completeness,
                    late_count,
                    no_completed_count,
                    no_plan_count,
                    problems_count,
                    notes_count,
                    risk,
                    "; ".join(reasons) if reasons else "Системных проблем не выявлено",
                ]
            )

        rows.sort(
            key=lambda row: (
                row[0],
                row[4],
                row[5],
            )
        )

        sheet = workbook.create_sheet("Author Review")

        sheet.append(
            [
                "Приоритет",
                "Автор",
                "Роль",
                "Отчетов",
                "Среднее качество",
                "Средняя полнота",
                "Опозданий",
                "Без выполненных задач",
                "Без плана",
                "С проблемами",
                "С замечаниями",
                "Риск",
                "Причины",
            ]
        )

        priority_labels = {
            1: "Высокий",
            2: "Средний",
            3: "Низкий",
        }

        for row in rows:
            row[0] = priority_labels.get(
                row[0],
                "Низкий",
            )
            sheet.append(row)

        self.format_sheet(sheet)

    def create_daily_review_sheet(
        self,
        workbook,
        structured_reports: list[StructuredDailyReport],
        report_metrics: list[DailyReportMetrics],
    ):
        """Создает лист с daily-отчетами, требующими внимания лида."""

        metrics_by_key = {
            (metric.author, metric.report_date): metric
            for metric in report_metrics
        }

        rows = []

        for report in structured_reports:
            metric = metrics_by_key.get(
                (
                    report.author,
                    report.report_date,
                )
            )

            reasons = []
            priority = 3

            if metric is None:
                reasons.append("Нет рассчитанных метрик")
                priority = 1
                quality_score = None
                completeness_score = None
                is_late = "Нет"
                notes = ""
            else:
                quality_score = metric.quality_score
                completeness_score = metric.completeness_score
                is_late = "Да" if metric.is_late else "Нет"
                notes = "; ".join(metric.notes)

                if metric.quality_score < 70:
                    reasons.append("Низкое качество отчета")
                    priority = min(priority, 1)

                if metric.completeness_score < 80:
                    reasons.append("Низкая полнота отчета")
                    priority = min(priority, 2)

                if metric.is_late:
                    reasons.append("Отчет опубликован позже нормы")
                    priority = min(priority, 2)

                if metric.completed_tasks_count == 0:
                    reasons.append("Нет выполненных задач за вчера")
                    priority = min(priority, 1)

                if metric.planned_tasks_count == 0:
                    reasons.append("Нет плана на сегодня")
                    priority = min(priority, 1)

                if metric.problems_count > 0:
                    reasons.append("Есть проблемы или блокеры")
                    priority = min(priority, 2)

                if metric.notes:
                    reasons.extend(metric.notes)

            if not reasons:
                continue

            unique_reasons = list(dict.fromkeys(reasons))

            rows.append(
                [
                    priority,
                    self.excel_safe_value(report.report_date),
                    report.author,
                    report.role,
                    quality_score,
                    completeness_score,
                    is_late,
                    "; ".join(unique_reasons),
                    report.raw_text,
                    notes,
                ]
            )

        rows.sort(
            key=lambda row: (
                row[0],
                row[4] if row[4] is not None else 999,
                row[5] if row[5] is not None else 999,
            )
        )

        sheet = workbook.create_sheet("Daily Review")

        sheet.append(
            [
                "Приоритет",
                "Дата",
                "Автор",
                "Роль",
                "Качество",
                "Полнота",
                "Опоздал",
                "Причина проверки",
                "Текст отчета",
                "Замечания",
            ]
        )

        priority_labels = {
            1: "Высокий",
            2: "Средний",
            3: "Низкий",
        }

        for row in rows:
            row[0] = priority_labels.get(
                row[0],
                "Низкий",
            )
            sheet.append(row)

        self.format_sheet(sheet)

    def excel_safe_value(self, value):
        """Преобразует значения в формат, безопасный для Excel."""

        if isinstance(value, datetime):
            return value.replace(tzinfo=None)

        if isinstance(value, date):
            return value.isoformat()

        if isinstance(value, time):
            return value.strftime("%H:%M")

        return value

    def format_sheet(self, sheet):
        """Оформляет лист Excel."""

        for cell in sheet[1]:
            cell.fill = self.HEADER_FILL
            cell.font = self.HEADER_FONT
            cell.alignment = self.HEADER_ALIGNMENT

        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions

        for column in sheet.columns:
            max_length = max(
                len(str(cell.value)) if cell.value else 0
                for cell in column
            )

            sheet.column_dimensions[get_column_letter(column[0].column)].width = min(
                max_length + 3,
                80,
            )
